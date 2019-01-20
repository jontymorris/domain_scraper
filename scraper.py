#!/bin/python2

import os
from sys import argv
from search import Searcher
from worker import Worker
from openpyxl import load_workbook, Workbook

def show_args():
    """ Prints the required args and closes the program """

    print "\nUsage: theharvester options"
    print "\t-d: Domain to search or company name"
    print "\t-z: Filename containing domains to search (one domain per line)"
    print "\t-f: save the results into an txt file"
    print "\t-l: limit the number of results to work with"

    exit(0)

def load_domains(path):
    """ Retrives a list of domains from the file """

    file = open(path, "r")
    domains = file.read().split("\n")

    for i in range(len(domains)):
        domains[i] = domains[i].lower()

        for item in ["http://", "www.", "https://"]:
            domains[i] = domains[i].replace(item, "")
    
    domains = list(set(domains))

    return domains

def write_excel_file(output_file, workers):
    """ write all of the emails to an excel file """

    output_file = output_file.replace('txt', 'xlsx')
    
    wb = Workbook()

    # check if there is an existing excel file
    already_exists = os.path.exists(output_file)
    existing_emails = []

    # if already existing, then load it
    if already_exists:
        wb = load_workbook(filename = output_file)
    
    ws = wb.active
    offset = 1

    # read the existing emails and get last offset
    if already_exists:
        while True:
            email = ws['B'+str(offset)].value

            if email == None:
                break
            
            existing_emails.append(email)
            offset += 1

    # loop through all the workers
    for worker in workers:
        if len(worker.emails) == 0:
            continue
        
        # loop through emails in the worker
        for email in worker.emails:
            # is this email already in the file?
            if email in existing_emails:
                continue

            # write email to file
            ws['A'+str(offset)] = worker.domain
            ws['B'+str(offset)] = email

            offset += 1

    # save and close the excel file
    wb.save(output_file)
    wb.close()

def main(args):
    searcher = Searcher(args["limit"])
    workers = []

    if "domains" in args:
        # load domains from file
        domains = load_domains(args["domains"])
        
        for domain in domains:
            if domain == "":
                continue
            
            # lookup in search engine
            result = searcher.google_search(domain)

            # start the worker
            worker = Worker(domain, result.urls, result.page_source)
            workers.append(worker)
        
        print "\nNow waiting for workers to finish"
    
    else:
        # lookup in search engine
        result = searcher.google_search(args["domain"])

        # start the worker
        worker = Worker(args["domain"], result.urls, result.page_source)
        workers.append(worker)
    
    searcher.close()

    # wait for all workers to finish
    for worker in workers:
        worker.wait()

    # write emails to a file
    if "output" in args:
        write_excel_file(args["output"], workers)

    print "\nFinished scraping!\n"

    # output all emails
    for worker in workers:
        for email in worker.emails:
            print "> " + email

if __name__ == "__main__":
    argv.pop(0) # remove the default argument

    # Check if enough args were passed in
    if len(argv) < 2 or len(argv) % 2 != 0:
        print "Not enough args provided!"
        show_args()
    
    args = {}

    # Pass the arguments
    index = 0
    while index < len(argv):
        try:
            if argv[index] == "-d":
                args["domain"] = argv[index+1]
            
            elif argv[index] == "-z":
                args["domains"] = argv[index+1]
            
            elif argv[index] == "-f":
                args["output"] = argv[index+1]
            
            elif argv[index] == "-l":
                args["limit"] = int(argv[index+1])
            
            else:
                print "Invalid argument!"
                print "\t" + argv[index]
                show_args()
            
            index += 2
        except:
            print "Error while parsing " + argv[index]
            show_args()

    # ensure a domain is provided
    if "domain" not in args and "domains" not in args:
        print "Must provide the domain to search!"
        show_args()

    # adjust limit
    if "limit" not in args:
        args["limit"] = 40
    elif args["limit"] > 60:
        args["limit"] = 60
        print "Lowing the search limit to 60"
    
    main(args)